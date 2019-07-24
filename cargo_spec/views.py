from django.contrib.auth.models import User
from django.shortcuts import render, redirect, get_object_or_404, HttpResponseRedirect
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.db import transaction
from django.contrib.auth.mixins import LoginRequiredMixin
from datetime import datetime
from django.http import HttpResponse
# from django.http import HttpResponseForbidden
from django.core.exceptions import PermissionDenied

from django.urls import resolve

from .models import Specification, CargoContent, SpecificationDocument
from .forms import SpecificationForm, CargoContentForm, CargoContentFormSet, SingleSpecificationDocumentForm, MultipleSpecificationDocumentForm
from .utils import specification_marking, check_scan_file

from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import pandas as pd

class MyListView(LoginRequiredMixin, ListView):
    model = Specification
    template_name = 'my_lists.html'
    ordering = ['-marking']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        owner_ = self.request.user
        return queryset.filter(owner = owner_)


class AddSpecificationView(LoginRequiredMixin, CreateView):
    model = Specification
    template_name = 'specification_form.html'
    form_class = SpecificationForm
    
    def get_context_data(self, **kwargs):
        data = super(AddSpecificationView, self).get_context_data(**kwargs)
        if self.request.POST:
            data['cargos'] = CargoContentFormSet(self.request.POST)
        else:
            data['cargos'] = CargoContentFormSet()
        return data
    
    def form_valid(self, form):
        context = self.get_context_data()
        cargos = context['cargos']
        with transaction.atomic():
            form.instance.owner = self.request.user
            form.instance.marking = specification_marking(form)
            
            self.object = form.save()

            # Sprawdza czy wszystkie wymagane pola dla modelu CargoContent są wypełnione, jeśli nie, to zwraca formularz
            if cargos.is_valid() == False:
                return self.render_to_response(self.get_context_data(form=form, cargos=cargos))

            if cargos.is_valid():
                cargos.instance = self.object
                cargos.save()
        return super(AddSpecificationView, self).form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('cargo_spec:my-lists')
    
    
class ModifySpecificationView(LoginRequiredMixin, UpdateView):
    model = Specification
    template_name = 'specification_form.html'
    form_class = SpecificationForm
    
    def get_context_data(self, **kwargs):
        data = super(ModifySpecificationView, self).get_context_data(**kwargs)
        
        if (self.object.owner != self.request.user) or (self.object.approved is True):
            raise PermissionDenied()
            
        if self.request.POST:
            data['cargos'] = CargoContentFormSet(self.request.POST, instance=self.object)
        else:
            data['cargos'] = CargoContentFormSet(instance=self.object)
        return data
    
    def form_valid(self, form):
        context = self.get_context_data()
        cargos = context['cargos']
        with transaction.atomic():
            self.object = form.save()
            
            # Sprawdza czy wszystkie wymagane pola dla modelu CargoContent są wypełnione, jeśli nie, to zwraca formularz
            if cargos.is_valid() == False:
                return self.render_to_response(self.get_context_data(form=form, cargos=cargos))
            
            if cargos.is_valid():
                cargos.instance = self.object
                cargos.save()
        return super(ModifySpecificationView, self).form_valid(form)
        
    def get_success_url(self):
        return reverse_lazy('cargo_spec:my-lists')
    

class AcceptSpecificationView(LoginRequiredMixin, View):
    def post(self, request, pk):
        spec = get_object_or_404(Specification, pk=pk)
        spec.approved = True
        spec.save()
        
        return redirect('cargo_spec:spec-detail', spec.pk)
    
    
class DeleteSpecificationView(LoginRequiredMixin, DeleteView):
    model = Specification
    
    def get_success_url(self):
        return reverse_lazy('cargo_spec:my-lists')
    
    
class SpecificationDetailView(LoginRequiredMixin, DetailView):
    model = Specification
    template_name = 'specification_detail.html'
    
    def get(self, request, *args, **kwargs):
        context = super(SpecificationDetailView, self).get(request, *args, **kwargs)
        if (self.object.owner != self.request.user) or (self.object.approved is False):
            # return HttpResponseForbidden('Nie masz dostępu do specyfikacji innych użytkowników')
            raise PermissionDenied()
        return context
    def post(self,request,*args, **kwargs):
        context = super(SpecificationDetailView, self).get(request, *args, **kwargs)
        request.session['doc_id']=self.object.pk
        return redirect('cargo_spec:test')
        
        



class test(View,LoginRequiredMixin):
    def __init__(self):
        self.thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    def get(self,request):
        user=request.user
        spec=Specification.objects.get(pk=request.session['doc_id'])
        cargo=CargoContent.objects.filter(specification=spec)
        dataframe=pd.DataFrame(list(cargo.values()),index=[i for i in range(1,len(cargo.values())+1) ])
        wb=load_workbook('spis.xlsx')
        wb2=load_workbook('stopka.xlsx')
        ws=wb.active
        ws2=wb2.active
        print(request.session['doc_id'])
        ###################Dane
        ws['C4'].value=f'{user.first_name} {user.last_name}'
        ws['C5'].value= spec.package_type
        ws['C6'].value=f'{spec.dimension_length}/{spec.dimension_width}/{spec.dimension_height}' 
        ws['C7'].value=spec.weight
        ws['C8'].value=spec.storage

        
        
        footer=[ws2.cell(row=i,column=1).value for i in range(1,8)]
        actual_row=14
        ws.insert_rows(actual_row, amount=len(dataframe.values))
        for i in range(1,len(dataframe.values)+1):
            # print(dataframe.iloc[i])
            ws['A{}'.format(actual_row)]=i
            ws['B{}'.format(actual_row)]='{} ({})'.format(dataframe.loc[i].at['name'],dataframe.loc[i].at['serial_number'])  
            ws['C{}'.format(actual_row)]=dataframe.loc[i].at['quantity']
            ws.merge_cells(f'D{actual_row}:E{actual_row}')
            ws['D{}'.format(actual_row)]=dataframe.loc[i].at['unit_of_measurement']
            ws.merge_cells(f'F{actual_row}:G{actual_row}')
            ws['F{}'.format(actual_row)]=dataframe.loc[i].at['value']
            for item in list('ABCDEFG'):
                ws['{}{}'.format(item,actual_row)].border=self.thin_border

            actual_row+=1
        actual_row+=1
        for i in footer:
            actual_cell=ws[f'A{actual_row}']
            ws.merge_cells(f'A{actual_row}:G{actual_row}') 
            actual_cell.value=i
            alignment_obj = actual_cell.alignment.copy(horizontal='center', vertical='center')
            actual_cell.alignment = alignment_obj
            actual_row+=1
        wb.save('spis_1.xlsx') 
        return HttpResponse(user.username)


class SpecificationScanUploadView(LoginRequiredMixin, View):
    template_name = 'file_upload.html'
    form_class = SingleSpecificationDocumentForm

    def get(self, request, *args, **kwargs):
        spec = Specification.objects.get(pk=self.kwargs['pk'])
        if (spec.owner != self.request.user) or (check_scan_file(spec.marking)):
            raise PermissionDenied()
        form = self.form_class()
        return render(request, self.template_name, {'form': form})
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        if form.is_valid():
            spec = get_object_or_404(Specification, pk=self.kwargs['pk'])
            form.instance.file_type = resolve(request.path_info).url_name
            form.instance.specification = spec
            owner = self.request.user
            form.instance.uploaded_by = owner
            form.save()
            return redirect('cargo_spec:spec-detail', spec.pk)
        else:
            return render(request, self.template_name, {'form': form})
            
            
class SpecificationPhotoUploadView(LoginRequiredMixin, View):
    template_name = 'file_upload.html'
    form_class = MultipleSpecificationDocumentForm

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})
    
    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)
        files = request.FILES.getlist('file_field')
        if form.is_valid():
            spec = get_object_or_404(Specification, pk=self.kwargs['pk'])
            owner = self.request.user
            for doc in files:
                SpecificationDocument.objects.create(
                    description = form.cleaned_data['description'],
                    document = doc,
                    uploaded_by = owner,
                    specification = spec
                )
            return redirect('cargo_spec:spec-detail', self.kwargs['pk'])
        else:
            return render(request, self.template_name, {'form': form})
            
            
class DeleteSpecificationDocumentView(LoginRequiredMixin, DeleteView):
    model = SpecificationDocument
    
    def get_success_url(self):
        spk = self.kwargs['spk']
        return reverse_lazy('cargo_spec:spec-detail', kwargs={'pk': spk})
