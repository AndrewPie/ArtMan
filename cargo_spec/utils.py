import os
from datetime import datetime
from django.conf import settings

from .models import Specification,CargoContent,SpecificationDocumentsExcel
import os
from django.core.files import File
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import pandas as pd

def specification_marking(self):
    mark = f'{datetime.now().year}-{self.instance.owner.first_name[0]}{self.instance.owner.last_name[0]}-{self.instance.package_type[:2].upper()}'
    
    query_list = Specification.objects.filter(marking__contains=mark).values_list('marking', flat=True)
    
    try:
        val_list = [int(i[11:]) for i in query_list]
        number = min(set(range(1, len(query_list) + 1)) - set(val_list))
    except Exception:
        number = len(query_list) + 1
    
    return f'{mark}-{number}'


def check_scan_file(marking):
    scan_dir = f'{settings.BASE_DIR}/media/cargo_spec/{marking}/scan'
    if os.path.exists(scan_dir) and os.path.isdir(scan_dir):
        if not os.listdir(scan_dir):
            return False
        else:    
            return True
    else:
        return False


def AddExcelToSpec(request,specification):
    path_ti_app_dir=os.path.dirname(os.path.realpath(__file__))
    path_project_dir=os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
    print(path_ti_app_dir)
    thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
    user=request.user
    # spec=Specification.objects.get(pk=request.session['doc_id'])
    spec=specification
    cargo=CargoContent.objects.filter(specification=spec)
    dataframe=pd.DataFrame(list(cargo.values()),index=[i for i in range(1,len(cargo.values())+1) ])
    wb=load_workbook(os.path.join(path_ti_app_dir,'static/cargo_raport/spis.xlsx'))
    wb2=load_workbook(os.path.join(path_ti_app_dir,'static/cargo_raport/stopka.xlsx'))
    pdf_out_path=(os.path.join(path_ti_app_dir,'/media/cargo_spec/pdf_files/'))
    ws=wb.active
    ws2=wb2.active
    
    ###################Dane
    ws['C4'].value=f'{user.first_name} {user.last_name}'
    ws['C5'].value= spec.package_type
    ws['C6'].value=f'{spec.dimension_length}/{spec.dimension_width}/{spec.dimension_height}' 
    ws['C7'].value=spec.weight
    ws['C8'].value=spec.storage
    
    
    footer=[ws2.cell(row=i,column=1).value for i in range(1,8)]
    actual_row=14
    ws.insert_rows(actual_row, amount=len(dataframe.values))
    for i in range(1,len(dataframe.values)+1):
        # print(dataframe.iloc[i])
        ws['A{}'.format(actual_row)]=i
        ws['B{}'.format(actual_row)]='{} ({})'.format(dataframe.loc[i].at['name'],dataframe.loc[i].at['serial_number'])  
        ws['C{}'.format(actual_row)]=dataframe.loc[i].at['quantity']
        ws.merge_cells(f'D{actual_row}:E{actual_row}')
        ws['D{}'.format(actual_row)]=dataframe.loc[i].at['unit_of_measurement']
        ws.merge_cells(f'F{actual_row}:G{actual_row}')
        ws['F{}'.format(actual_row)]=dataframe.loc[i].at['value']
        for item in list('ABCDEFG'):
            ws['{}{}'.format(item,actual_row)].border=thin_border
        actual_row+=1
    actual_row+=1
    
    temp_file=f'{path_project_dir}/media/cargo_spec/excel_files/temp.xlsx'
    # file_path=f'{path_project_dir}/media/cargo_spec/excel_files/{spec.marking}.xlsx'
    for i in footer:
        actual_cell=ws[f'A{actual_row}']
        ws.merge_cells(f'A{actual_row}:G{actual_row}') 
        actual_cell.value=i
        alignment_obj = actual_cell.alignment.copy(horizontal='center', vertical='center')
        actual_cell.alignment = alignment_obj
        actual_row+=1
    wb.save(temp_file)
    model=SpecificationDocumentsExcel(spec=specification)
    model.excel.save(f'{spec.marking}.xlsx',File(open(temp_file,'rb')))
    os.remove(temp_file)

